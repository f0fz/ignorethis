#!/usr/bin/env python
# -*- coding: utf-8 -*-

import logging
from openpyxl import load_workbook

from main import rList

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)


def dumpExcel(ws):
    row_number = 2
    while ws['A' + str(row_number)].value is not None:
        NRIC = ws['A' + str(row_number)].value
        GRP_ID = ws['B' + str(row_number)].value
        PERSON.append({'NRIC': NRIC,
                       'GRP1': GRP_ID,
                       'GRP1_REG': ''})
        rList.mset({NRIC: ''})  # dump NRIC into redis
        row_number += 1


def findNRIC(mainlist, nric):
    listOfNRICIndices = []
    for i in range(0,len(mainlist)):
        if nric.lower() == mainlist[i]['NRIC'].lower():
            listOfNRICIndices.append(i)

    if len(listOfNRICIndices) == 1:
        return listOfNRICIndices[0]
    else:
        return None


def returnSeating(mainlist, nric):
    index = findNRIC(mainlist, nric)
    if index is not None:
        mainlist[index]['GRP1_REG'] = 'P'
        rList.set(mainlist[index]['NRIC'], 'P')  # mark attendance in rList
        return mainlist[index]['GRP1']

    return None


def createFile():
    logger.info('Good day, admin. Loading excel file into memory and updating...')

    main_workbook = load_workbook('SeminarDatasheet.xlsx')
    worksheet = main_workbook['Sheet1']
    row_number = 2
    cellNRIC = 'A' + str(row_number)

    while worksheet[cellNRIC].value is not None:
        NRIC = worksheet[cellNRIC].value
        if rList.get(NRIC).decode('utf-8') == 'P':
            worksheet['C' + str(row_number)] = 'P'
        row_number += 1
        cellNRIC = 'A' + str(row_number)

    logger.info('Updating complete, saving excel file')
    main_workbook.save('Attendance.xlsx')
    logger.info("File creation complete")


# no longer needed due to redis

# def saveFile(mainlist, worksheet, main_workbook):
#     for i in range(0, len(mainlist)):
#         row_number = str(i + 2)
#         worksheet['C' + str(row_number)].value = mainlist[i]['GRP1_REG']
#     logger.info("saveFile is run")
#     main_workbook.save('SeminarDatasheet.xlsx')
#     # should only be run at end of bot life
